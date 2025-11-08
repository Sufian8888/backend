#!/usr/bin/env python3
"""
Script to create Excel file with test login credentials for PneuShop application
"""

import os
import django
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'pneushop.settings')
django.setup()

from accounts.models import CustomUser

def create_credentials_excel():
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Credentials"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    data_font = Font(name="Arial", size=11)
    data_alignment = Alignment(horizontal="left", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Account Type",
        "Name", 
        "Email",
        "Username",
        "Password",
        "Role",
        "Status",
        "Notes"
    ]
    
    # Add headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Test credentials data
    credentials_data = [
        [
            "Admin Account",
            "Admin User",
            "admin@pneushop.tn",
            "admin",
            "AdminSecure2024!",
            "Admin",
            "Active",
            "Main admin account for testing"
        ],
        [
            "Test Registration",
            "Test User",
            "testuser@pneushop.tn",
            "test_user",
            "TestSecure2024!",
            "Customer",
            "For Registration",
            "Use this for testing registration form"
        ],
        [
            "Demo Customer",
            "Demo Customer",
            "customer@pneushop.tn",
            "demo_customer",
            "CustomerSecure2024!",
            "Customer",
            "For Testing",
            "General customer account for testing"
        ]
    ]
    
    # Add data rows
    for row_num, row_data in enumerate(credentials_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
            
            # Color coding
            if row_data[0] == "Admin Account":
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            elif row_data[0] == "Test Registration":
                cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="F0F8E6", end_color="F0F8E6", fill_type="solid")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add metadata sheet
    meta_ws = wb.create_sheet(title="Info")
    
    meta_info = [
        ["PneuShop Test Credentials", ""],
        ["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["Frontend URLs:", ""],
        ["Login Page:", "http://localhost:3000/auth/login"],
        ["Register Page:", "http://localhost:3000/auth/register"],
        ["Debug Page:", "http://localhost:3000/debug"],
        ["", ""],
        ["Backend URLs:", ""],
        ["Django Admin:", "http://localhost:8000/admin/"],
        ["API Base:", "http://localhost:8000/api/"],
        ["", ""],
        ["Password Requirements:", ""],
        ["- Minimum 8 characters", ""],
        ["- Not too common (avoid 'password123')", ""],
        ["- Mix of letters, numbers, symbols", ""],
        ["- Not similar to username/email", ""],
        ["", ""],
        ["Quick Test Steps:", ""],
        ["1. Start Django server: python manage.py runserver", ""],
        ["2. Start Next.js server: npm run dev", ""],
        ["3. Use quick-fill buttons in forms", ""],
        ["4. Check browser console for debug info", ""]
    ]
    
    for row_num, (key, value) in enumerate(meta_info, 1):
        meta_ws.cell(row=row_num, column=1, value=key).font = Font(bold=True if value else False)
        meta_ws.cell(row=row_num, column=2, value=value)
    
    # Auto-adjust meta sheet columns
    for column in meta_ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        meta_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save file
    filename = "PneuShop_Test_Credentials.xlsx"
    wb.save(filename)
    print(f"✅ Excel file created: {filename}")
    
    # Also create a simple CSV for backup
    import csv
    csv_filename = "PneuShop_Test_Credentials.csv"
    with open(csv_filename, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(headers)
        writer.writerows(credentials_data)
    print(f"✅ CSV backup created: {csv_filename}")
    
    return filename, csv_filename

if __name__ == "__main__":
    try:
        excel_file, csv_file = create_credentials_excel()
        print(f"\n📄 Files created in: {os.getcwd()}")
        print(f"📊 Excel: {excel_file}")
        print(f"📋 CSV: {csv_file}")
        
        # Show current users in database
        print(f"\n👥 Current users in database:")
        users = CustomUser.objects.all()
        for user in users:
            print(f"  - {user.email} ({user.username}) - {'Admin' if user.is_staff else 'Customer'}")
            
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()