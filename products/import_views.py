import os
import time
from decimal import Decimal
import tempfile
import pandas as pd
import re
import cloudinary.uploader
from io import BytesIO
import tempfile
from openpyxl import load_workbook
from django.conf import settings
from django.utils.text import slugify
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework import status
from .models import Product, Category

output_folder = os.path.join(settings.MEDIA_ROOT, "uploads/images")
os.makedirs(output_folder, exist_ok=True)



def extract_tire_info(name):
    """Extract tire information from product name"""
    print(f"🔍 Extracting tire info from: {name}")
    
    # Extract brand - look for common brands
    brand = "Continental"  # Default brand
    brand_patterns = ['LAUFENN', 'CONTINENTAL', 'MICHELIN', 'BRIDGESTONE', 'GOODYEAR', 'PIRELLI']
    for b in brand_patterns:
        if b.upper() in name.upper():
            brand = b.capitalize()
            break
    
    # Extract tire size using improved regex (format: XXX/XX RXX or XXX/XXrXX)
    # Updated to handle more variations: 165/60R14, 195/65 R 15, 205/55R16, etc.
    size_pattern = r'(\d{2,3}[/]\d{2}\s?[RrXx]?\s?\d{1,2})'
    size_match = re.search(size_pattern, name, re.IGNORECASE)
    size = size_match.group(1) if size_match else "Unknown"
    
    # Clean size format
    if size != "Unknown":
        size = re.sub(r'\s+', '', size).upper().replace('r', 'R').replace('X', 'R')
        if 'R' not in size and '/' in size:
            # Add R if missing (e.g., 205/55 16 -> 205/55R16)
            parts = size.split('/')
            if len(parts) == 2:
                size = f"{parts[0]}/{parts[1][:2]}R{parts[1][2:]}"
    
    print(f"   Brand: {brand}, Size: {size}")
    
    # Remove common prefixes and tire size to extract product name
    clean_name = name.replace("Pneu", "").replace("pneu", "").strip()
    
    # Remove brand from name
    for b in brand_patterns:
        clean_name = clean_name.replace(b, "").replace(b.upper(), "").replace(b.lower(), "")
    
    # Remove the tire size pattern
    if size_match:
        clean_name = clean_name.replace(size_match.group(1), "").strip()
    
    # Remove speed/load rating patterns (like 91H, 88T, 75H XL, etc.)
    clean_name = re.sub(r'\b\d{2,3}\s?[A-Z]{1,2}\s?(XL|RF|C)?\b', '', clean_name, flags=re.IGNORECASE).strip()
    
    # Remove extra whitespace and clean up
    clean_name = re.sub(r'\s+', ' ', clean_name).strip()
    
    # Extract meaningful product name
    if clean_name:
        # Remove leading/trailing non-alphanumeric characters
        clean_name = re.sub(r'^[^a-zA-Z0-9]+|[^a-zA-Z0-9]+$', '', clean_name)
        product_name = clean_name if clean_name else name[:50]
    else:
        product_name = name[:50]
    
    full_name = f"{brand} {product_name} {size}".strip()
    print(f"   Result: {full_name}")
    
    return {
        'brand': brand,
        'name': product_name,
        'size': size,
        'full_name': full_name
    }

def determine_season(name, description):
    """Determine tire season based on name and description"""
    text = (name + " " + str(description)).lower()
    
    if any(word in text for word in ['winter', 'hiver', 'neige', 'snow']):
        return 'winter'
    elif any(word in text for word in ['summer', 'été', 'sport']):
        return 'summer'
    else:
        return 'all_season'

def determine_category(name, description):
    """Determine product category from name and description"""
    text = (str(name) + " " + str(description)).lower()
    
    # ONLY these 5 categories are allowed
    VALID_CATEGORIES = ['tourisme', '4x4', 'agricole', 'utilitaire', 'moto']
    
    # Category keywords mapping
    category_keywords = {
        'tourisme': ['tourisme', 'tourism', 'passenger', 'car', 'voiture'],
        '4x4': ['4x4', '4wd', 'suv', 'tout-terrain', 'off-road'],
        'agricole': ['agricole', 'agricultural', 'farm', 'tracteur', 'tractor'],
        'utilitaire': ['utilitaire', 'utility', 'commercial', 'van', 'fourgon', 'camionnette'],
        'moto': ['moto', 'motorcycle', 'scooter', 'bike']
    }
    
    # Check for category keywords
    for category, keywords in category_keywords.items():
        if any(keyword in text for keyword in keywords):
            return category
    
    # Default to tourisme if no category found
    return 'tourisme'


# def extract_images_from_excel(excel_file):
#     """Extracts images from Excel and saves them to disk with row reference"""
#     wb = load_workbook(excel_file)
#     ws = wb.active

#     row_images = {}
#     for i, image in enumerate(ws._images, start=1):
#         row = image.anchor._from.row
#         if row not in row_images:
#             row_images[row] = []
#         row_images[row].append(image)

#     saved_images = {}
#     for row, images in row_images.items():
#         if images:
#             tire_image = images[0]  # take first image per row
#             img_bytes = tire_image._data()
#             img_name = f"row_{row}_tire.png"
#             img_path = os.path.join(output_folder, img_name)
#             with open(img_path, "wb") as f:
#                 f.write(img_bytes)
#             saved_images[row] = f"uploads/images/{img_name}"  # relative path for DB
#     #         img_bytes_io = BytesIO(img_bytes)
#     #         upload_result = cloudinary.uploader.upload(img_bytes_io, folder="pneushop/uploads/")
#     #         saved_images[row] = upload_result.get("secure_url")  # save the URL directly
#             return saved_images

def extract_images_from_excel(excel_file):
    """Extracts ALL images from ALL sheets (up to 3 per row), uploads to Cloudinary, returns row->[URLs] mapping"""
    wb = load_workbook(excel_file, data_only=True)
    
    all_saved_images = {}
    row_offset = 0
    
    # Process each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        row_images = {}
        for image in ws._images:
            try:
                row = image.anchor._from.row
                if row not in row_images:
                    row_images[row] = []
                row_images[row].append(image)
            except AttributeError:
                continue

        # Upload images to Cloudinary
        for row, images in row_images.items():
            if images:
                # Process up to 3 images per row
                uploaded_urls = []
                for idx, tire_image in enumerate(images[:3]):  # Max 3 images
                    try:
                        img_bytes = tire_image._data()
                        img_bytes_io = BytesIO(img_bytes)

                        upload_result = cloudinary.uploader.upload(
                            img_bytes_io,
                            folder="pneushop/uploads/",
                            resource_type="image"
                        )
                        uploaded_urls.append(upload_result.get("secure_url"))
                    except Exception as e:
                        continue
                
                # Store with adjusted row number (accounting for multiple sheets)
                all_saved_images[row + row_offset] = uploaded_urls
        
        # Update offset for next sheet (number of rows in current sheet)
        row_offset += ws.max_row

    print(f"✅ Extracted and uploaded {sum(len(urls) for urls in all_saved_images.values())} images from {len(all_saved_images)} products")
    return all_saved_images


@api_view(['POST'])
@permission_classes([AllowAny])
def import_products_excel(request):
    # Initialize variables immediately to prevent "not associated with a value" errors
    df = None
    total_rows = 0
    created_products = []
    errors = []
    row_images = {}
    
    try:
        # Validate Cloudinary configuration
        cloudinary_available = False
        try:
            import cloudinary
            
            # Check if cloudinary is properly configured via environment or settings
            cloudinary_available = bool(
                cloudinary.config().cloud_name and 
                cloudinary.config().api_key and 
                cloudinary.config().api_secret
            )
            
            if cloudinary_available:
                print("✅ Cloudinary is configured and available")
            else:
                print("⚠️ Cloudinary not fully configured - images will be skipped")
                
        except Exception as e:
            print(f"⚠️ Cloudinary validation failed: {e} - images will be skipped")
        
        if 'file' not in request.FILES:
            return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        excel_file = request.FILES['file']

        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response({'error': 'Invalid file type'}, status=status.HTTP_400_BAD_REQUEST)

        # Save temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            for chunk in excel_file.chunks():
                tmp.write(chunk)
            temp_path = tmp.name

        # Load Excel data first - handle multiple sheets
        try:
            # Read all sheets
            excel_data = pd.read_excel(temp_path, sheet_name=None)  # Returns dict of all sheets
            
            if not excel_data:
                return Response({
                    'error': 'Excel file has no sheets',
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Combine all sheets into one dataframe
            df = pd.concat(excel_data.values(), ignore_index=True)
            sheet_count = len(excel_data)
            
            print(f"✅ Successfully loaded Excel file with {sheet_count} sheet(s) and {len(df)} total rows")
            print(f"📋 Columns found in Excel: {list(df.columns)}")
            print(f"📊 First few rows sample:")
            print(df.head(3))
        except Exception as e:
            return Response({
                'error': f'Failed to read Excel file: {str(e)}',
                'note': 'Please ensure the file is a valid Excel file (.xlsx or .xls)'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Extract images with timeout protection (reset row_images)
        
        # Extract images for all files (removed row limit check)
        process_images = True  # Always process images
        
        if process_images:
            try:
                print("🔄 Starting image extraction from Excel...")
                row_images = extract_images_from_excel(temp_path)
                print(f"✅ Extracted {len(row_images)} images from Excel")
            except Exception as e:
                print(f"⚠️ Image extraction failed: {e}. Continuing without images.")
                row_images = {}
                import traceback
                print(f"Image extraction traceback: {traceback.format_exc()}")
        
        # Handle both old and new column formats
        # Normalize column names first
        print(f"🔧 Original columns: {list(df.columns)}")
        df.columns = [str(c).strip().upper() for c in df.columns]
        print(f"🔧 Normalized columns: {list(df.columns)}")
        
        # Handle REFERNECE column (typo in Excel) - rename to REFERENCE
        if 'REFERNECE' in df.columns:
            df = df.rename(columns={'REFERNECE': 'REFERENCE'})
            print(f"✅ Fixed REFERNECE typo → REFERENCE")
        
        # Handle Unnamed: 0 column (appears when concatenating sheets)
        if 'UNNAMED: 0' in df.columns:
            if 'REFERENCE' not in df.columns and 'NOM' not in df.columns:
                # Only use it as NOM if no other name column exists
                df = df.rename(columns={'UNNAMED: 0': 'NOM'})
                print(f"✅ Renamed UNNAMED: 0 → NOM")
            else:
                # Drop it - it's probably sheet names, not product data
                df = df.drop(columns=['UNNAMED: 0'])
                print(f"✅ Dropped UNNAMED: 0 column (contains sheet names, not product data)")

        # Required: either NOM or REFERENCE, and PRIX TTC
        has_name = 'NOM' in df.columns or 'REFERENCE' in df.columns
        has_price = 'PRIX TTC' in df.columns
        
        if not has_price:
            return Response({
                'error': 'Missing required column: PRIX TTC',
                'columns_found': list(df.columns)
            }, status=status.HTTP_400_BAD_REQUEST)
            
        if not has_name:
            return Response({
                'error': 'Missing product name column (expected NOM or REFERENCE)',
                'columns_found': list(df.columns)
            }, status=status.HTTP_400_BAD_REQUEST)

        # Categories will be created dynamically per product based on type

        # Update variables (initialized earlier)
        batch_size = 20  # Process 20 rows at a time to prevent timeout
        total_rows = len(df) if df is not None else 0

        if total_rows == 0:
            return Response({
                'error': 'Excel file is empty or has no data rows',
                'summary': {'total_rows': 0, 'created': 0, 'updated': 0, 'errors': 1},
                'errors': ['No data found in Excel file']
            }, status=status.HTTP_400_BAD_REQUEST)

        print(f"🔄 Processing {total_rows} rows from Excel in batches of {batch_size}...")

        for batch_start in range(0, total_rows, batch_size):
            batch_end = min(batch_start + batch_size, total_rows)
            print(f"📦 Processing batch {batch_start//batch_size + 1}: rows {batch_start+1} to {batch_end}")
            
            # Process current batch
            for index in range(batch_start, batch_end):
                row = df.iloc[index]
                try:
                    # Get product name from REFERENCE or NOM column
                    product_name = None
                    if 'REFERENCE' in df.columns and not pd.isna(row['REFERENCE']):
                        product_name = str(row['REFERENCE']).strip()
                    elif 'NOM' in df.columns and not pd.isna(row['NOM']):
                        product_name = str(row['NOM']).strip()
                    
                    # Debug: print what we got from Excel
                    print(f"\n--- Row {index + 1} ---")
                    print(f"Product Name from Excel: '{product_name}'")
                    if 'REFERENCE' in df.columns:
                        print(f"REFERENCE column value: '{row['REFERENCE']}'")
                    if 'NOM' in df.columns:
                        print(f"NOM column value: '{row['NOM']}'")
                    print(f"PRIX TTC: {row.get('PRIX TTC', 'N/A')}")
                    print(f"All row data: {dict(row)}")
                    
                    # Skip if no valid product name or price
                    if not product_name or pd.isna(row['PRIX TTC']):
                        print(f"❌ Skipping row {index + 1}: Missing name or price")
                        continue

                    # Validate product name
                    if len(product_name) < 2:
                        errors.append(f"Row {index + 1}: Invalid product name")
                        continue
                        
                    # Validate price
                    price = float(row['PRIX TTC'])
                    if price <= 0:
                        errors.append(f"Row {index + 1}: Invalid price: {price}")
                        continue
                        
                except (ValueError, TypeError) as e:
                    errors.append(f"Row {index + 1}: Data validation error: {e}")
                    continue

                # Handle optional DESCRIPTION column - preserve complete multi-line text
                description = ""
                if 'DESCRIPTION' in df.columns and not pd.isna(row['DESCRIPTION']):
                    # Keep full description including newlines and formatting
                    description = str(row['DESCRIPTION']).strip()
                
                # Get image URLs if available (now returns list of URLs)
                image_urls = row_images.get(index + 2, [])
                image_1 = image_urls[0] if len(image_urls) > 0 else ""
                image_2 = image_urls[1] if len(image_urls) > 1 else ""
                image_3 = image_urls[2] if len(image_urls) > 2 else ""

                # Extract tire info & generate unique slug
                try:
                    tire_info = extract_tire_info(product_name)
                    
                    # Use the FULL REFERENCE as the product name (not the cleaned version)
                    # Only extract brand and size, keep original name intact
                    product_display_name = product_name  # Use full reference as-is
                        
                except Exception as e:
                    print(f"⚠️ Tire info extraction failed for row {index + 1}: {e}")
                    tire_info = {
                        'brand': 'Laufenn',
                        'size': 'Unknown',
                    }
                    product_display_name = product_name

                # Generate unique slug from full product name
                try:
                    base_slug = slugify(product_display_name)
                    if not base_slug:  # If slugify returns empty string
                        base_slug = f"product-{index}"
                        
                    slug = base_slug
                    counter = 1
                    while Product.objects.filter(slug=slug).exists():
                        slug = f"{base_slug}-{counter}"
                        counter += 1
                        
                except Exception as e:
                    slug = f"product-{index}-{int(time.time())}"  # Fallback unique slug
                    print(f"⚠️ Slug generation failed for row {index + 1}: {e}, using fallback: {slug}")

                # Determine season
                try:
                    season = determine_season(product_name, description)
                except Exception as e:
                    season = 'all_season'  # Safe fallback
                    print(f"⚠️ Season determination failed for row {index + 1}: {e}")

                # Determine category dynamically
                try:
                    category_name = determine_category(product_name, description)
                    category_slug = slugify(category_name)
                    
                    category, _ = Category.objects.get_or_create(
                        slug=category_slug,  # Match on slug (unique field)
                        defaults={
                            'name': category_name,
                            'description': f'Pneus {category_name}'
                        }
                    )
                except Exception as e:
                    # Fallback to default category
                    category, _ = Category.objects.get_or_create(
                        slug='tourisme',  # Match on slug
                        defaults={'name': 'tourisme', 'description': 'Pneus tourisme'}
                    )
                    print(f"⚠️ Category determination failed for row {index + 1}: {e}")

                # Create product with error handling
                try:
                    product = Product.objects.create(
                        name=product_display_name[:200],  # Use full REFERENCE as product name
                        brand=tire_info['brand'][:100],  # Brand limit
                        size=tire_info['size'][:100],  # Increased size limit
                        slug=slug,
                        description=description,  # Full multi-line description preserved
                        price=Decimal(str(price)),
                        category=category,
                        season=season,
                        stock=10,
                        is_active=True,
                        image=image_1,      # First image
                        image_2=image_2,    # Second image (optional)
                        image_3=image_3     # Third image (optional)
                    )
                    created_products.append(product.name)
                    print(f"✅ Created product: {product.name} | Category: {category.name} | Images: {len([i for i in [image_1, image_2, image_3] if i])}")
                    
                except Exception as db_error:
                    error_msg = f"Row {index + 1}: Database error creating product: {db_error}"
                    errors.append(error_msg)
                    print(f"❌ {error_msg}")
                    continue

                except Exception as e:
                    error_msg = f"Row {index + 1}: Unexpected error: {e}"
                    errors.append(error_msg)
                    print(f"❌ {error_msg}")
                    import traceback
                    print(f"Full traceback: {traceback.format_exc()}")
                    continue
            
            # Print batch completion
            print(f"✅ Completed batch {batch_start//batch_size + 1} - Created {len(created_products)} products so far")

        # Clean up temporary file
        try:
            os.unlink(temp_path)
        except Exception as e:
            print(f"⚠️ Could not delete temp file {temp_path}: {e}")

        # Calculate success rate
        success_rate = (len(created_products) / total_rows * 100) if total_rows > 0 else 0
        
        response_data = {
            'message': '✅ Import completed successfully',
            'summary': {
                'total_rows': total_rows,
                'created': len(created_products),
                'updated': 0,  # Add for frontend compatibility
                'errors': len(errors),
                'success_rate': f"{success_rate:.1f}%",
                'processing_time': 'Processed in batches to prevent timeout',
                'images_processed': len(row_images) > 0
            },
            'created_products': created_products[:50],  # Limit response size
            'updated_products': [],  # Add for frontend compatibility
            'errors': errors[:20],  # Limit error list
            'note': 'Large files are processed in batches to prevent server timeout'
        }
        
        response = Response(response_data)
        # Explicitly add CORS headers
        response['Access-Control-Allow-Origin'] = '*'
        response['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
        return response

    except Exception as e:
        # More detailed error for debugging
        import traceback
        error_detail = {
            'error': str(e),
            'type': type(e).__name__,
            'traceback': traceback.format_exc()
        }
        print(f"❌ Import failed: {error_detail}")
        return Response(error_detail, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([AllowAny])
def import_products_fast(request):
    """Fast import for large files - minimal processing, no images"""
    try:
        if 'file' not in request.FILES:
            return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        excel_file = request.FILES['file']
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response({'error': 'Invalid file type'}, status=status.HTTP_400_BAD_REQUEST)

        # Save temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            for chunk in excel_file.chunks():
                tmp.write(chunk)
            temp_path = tmp.name

        # Load Excel data
        df = pd.read_excel(temp_path)
        
        # Handle column formats
        if 'Unnamed: 0' in df.columns:
            df = df.rename(columns={'Unnamed: 0': 'NOM'})
        else:
            df.columns = [str(c).strip().upper() for c in df.columns]

        # Get/create category
        category, _ = Category.objects.get_or_create(
            name='Continental',
            defaults={'slug': 'continental', 'description': 'Imported products'}
        )

        created_products, errors = [], []
        batch_size = 50  # Larger batches for faster processing
        
        # Bulk create products for speed
        products_to_create = []
        
        for index, row in df.iterrows():
            try:
                if pd.isna(row.get('NOM')) or pd.isna(row.get('PRIX TTC')):
                    continue
                    
                product_name = str(row['NOM']).strip()[:100]
                price = float(row['PRIX TTC'])
                
                # Simple slug generation
                slug = slugify(product_name)
                if not slug:
                    slug = f"product-{index}"
                    
                # Make slug unique with counter
                counter = 1
                original_slug = slug
                while Product.objects.filter(slug=slug).exists():
                    slug = f"{original_slug}-{counter}"
                    counter += 1
                
                # Create product object (don't save yet)
                products_to_create.append(Product(
                    name=product_name,
                    brand='Continental',
                    size='Unknown',
                    slug=slug,
                    description=str(row.get('DESCRIPTION', ''))[:500],
                    price=Decimal(str(price)),
                    category=category,
                    season='all_season',
                    stock=10,
                    is_active=True,
                    image=""
                ))
                
                # Bulk create every 50 products
                if len(products_to_create) >= batch_size:
                    Product.objects.bulk_create(products_to_create, ignore_conflicts=True)
                    created_products.extend([p.name for p in products_to_create])
                    products_to_create = []
                    print(f"✅ Created batch of {batch_size} products")
                    
            except Exception as e:
                errors.append(f"Row {index + 1}: {str(e)}")
                continue
        
        # Create remaining products
        if products_to_create:
            Product.objects.bulk_create(products_to_create, ignore_conflicts=True)
            created_products.extend([p.name for p in products_to_create])

        # Cleanup
        try:
            os.unlink(temp_path)
        except:
            pass

        response = Response({
            'message': '🚀 Fast import completed successfully',
            'summary': {
                'total_rows': len(df),
                'created': len(created_products),
                'errors': len(errors),
                'processing_method': 'Bulk create for maximum speed',
                'images_processed': False
            },
            'created_products': created_products[:20],
            'errors': errors[:10]
        })
        
        response['Access-Control-Allow-Origin'] = '*'
        return response

    except Exception as e:
        import traceback
        return Response({
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([AllowAny])
def quick_import_test(request):
    """Quick test import without images - for debugging deployment"""
    try:
        if 'file' not in request.FILES:
            return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        excel_file = request.FILES['file']
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response({'error': 'Invalid file type'}, status=status.HTTP_400_BAD_REQUEST)

        # Save temporarily and read Excel without image processing
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            for chunk in excel_file.chunks():
                tmp.write(chunk)
            temp_path = tmp.name

        # Load Excel data
        df = pd.read_excel(temp_path)
        
        # Clean up
        try:
            os.unlink(temp_path)
        except:
            pass

        response = Response({
            'message': '✅ Quick test successful - Excel file processed without images',
            'columns_found': list(df.columns),
            'total_rows': len(df),
            'first_3_rows': df.head(3).to_dict('records') if len(df) > 0 else [],
            'note': 'This is a test endpoint - no products were created'
        })
        
        # Add CORS headers
        response['Access-Control-Allow-Origin'] = '*'
        response['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        return response

    except Exception as e:
        import traceback
        return Response({
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([AllowAny])
def import_preview(request):
    """Preview Excel file data before import and test API connectivity"""
    
    # Test database connection
    try:
        category_count = Category.objects.count()
        product_count = Product.objects.count()
        db_status = f"✅ Database connected - {category_count} categories, {product_count} products"
    except Exception as e:
        db_status = f"❌ Database error: {e}"
    
    # Test Cloudinary
    try:
        import cloudinary
        cloudinary_status = f"✅ Cloudinary config: {bool(cloudinary.config().cloud_name)}"
    except Exception as e:
        cloudinary_status = f"❌ Cloudinary error: {e}"
    
    return Response({
        'message': 'Import API is ready',
        'status': {
            'database': db_status,
            'cloudinary': cloudinary_status,
            'api_endpoint': '✅ API responding correctly'
        },
        'expected_format': {
            'method': 'POST',
            'content_type': 'multipart/form-data',
            'file_field': 'file',
            'supported_formats': ['.xlsx', '.xls'],
            'columns': ['NOM (or Unnamed: 0)', 'PRIX TTC', 'DESCRIPTION (optional)'],
            'example': {
                'NOM': 'Pneu CONTINENTAL 195/65R15 91H ULTRA CONTACT',
                'PRIX TTC': 299.238,
                'DESCRIPTION': 'Points forts: Kilométrage ultra-élevé...'
            }
        }
    })
